import os
import json
import openpyxl
import tkinter as tk
import shutil

from tkinter import messagebox
from tkinter import filedialog
from PIL import Image, ImageTk
from pdf2image import convert_from_path
from alibabacloud_ocr_api20210707.client import Client as ocr_api20210707Client
from alibabacloud_tea_openapi import models as open_api_models
from alibabacloud_darabonba_stream.client import Client as StreamClient
from alibabacloud_ocr_api20210707 import models as ocr_api_20210707_models
from alibabacloud_tea_util import models as util_models


def invoice_to_excel(invoice_dic, file_path, sheet_name):
    excel_file_name = file_path[:-len("pdf")] + "xlsx"

    extracted_datas = []
    for item in invoice_dic["prism_keyValueInfo"]:
        if item["key"] == "taxClearanceDetails":
            for value in json.loads(item["value"]):
                extracted_data = {
                    "号码": invoice_dic.get("data").get("certificateNumber"),
                    "原凭证号": value.get("voucherNumber"),
                    "税种": value.get("taxType"),
                    "品目名称": value.get("itemName"),
                    "入库时间": value.get("date"),
                    "实缴金额": value.get("amount"),
                    "合计金额": invoice_dic.get("data").get("totalAmount")
                }
                extracted_datas.append(extracted_data)

    if len(extracted_datas) > 0:
        append_data_to_xlsx(excel_file_name, extracted_datas, sheet_name)


def append_data_to_xlsx(file_path, extracted_datas, sheet_name):
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook(filename=file_path)

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)

    headers = []
    if sheet.max_row == 1 and not any(cell.value for cell in sheet[1]):
        headers = list(extracted_datas[0].keys())
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)
    else:
        for cell in sheet[1]:
            headers.append(cell.value)

    for record in extracted_datas:
        row = [record.get(header) for header in headers]
        sheet.append(row)

    workbook.save(filename=file_path)


class OCR:
    def __init__(self, key, secret):
        config = open_api_models.Config(
            # 必填，请确保代码运行环境设置了环境变量 ALIBABA_CLOUD_ACCESS_KEY_ID。,
            access_key_id=key,
            # 必填，请确保代码运行环境设置了环境变量 ALIBABA_CLOUD_ACCESS_KEY_SECRET。,
            access_key_secret=secret
        )
        config.endpoint = f'ocr-api.cn-hangzhou.aliyuncs.com'
        self.client = ocr_api20210707Client(config)

    def get_client(self):
        return self.client

    def recognize(self, image_path, file_path, sheet_name) -> int:
        body_stream = StreamClient.read_from_file_path(image_path)
        recognize_request = ocr_api_20210707_models.RecognizeTaxClearanceCertificateRequest(body=body_stream)
        runtime = util_models.RuntimeOptions()
        try:
            resp = self.client.recognize_tax_clearance_certificate_with_options(recognize_request, runtime)
            response_dic = json.loads(resp.body.data)
            invoice_to_excel(response_dic, file_path, sheet_name)
        except Exception as e:
            directory = os.path.dirname(file_path)
            target_path = os.path.join(directory, os.path.basename(image_path))
            shutil.copy(image_path, target_path)
            return -1
        return 1


class Application(tk.Frame):
    def __init__(self, master=None, test=False):
        super().__init__(master)
        self.secret_file_path = None
        self.access_secret = None
        self.access_key = None
        self.test = test
        self.confirm_button = None
        self.upload_button = None
        self.access_key_entry = None
        self.access_secret_entry = None
        self.master = master
        self.pack()
        self.file_path = None
        self.create_buttons()
        self.doing = False
        self.failed = 0

    def create_buttons(self):
        canvas = tk.Canvas(self.master, width=800, height=600)
        canvas.pack()

        img = ImageTk.PhotoImage(Image.open("images/background.jpg").resize((800, 600), Image.ANTIALIAS))
        canvas.background = img  # Keep a reference in case this code is put in a function.
        canvas.create_image(0, 0, anchor=tk.NW, image=img)

        access_key_label = canvas.create_text(100, 30, text="Access Key:", fill="white")
        self.access_key_entry = tk.Entry(self.master, width=30)
        canvas.create_window(200, 75, window=self.access_key_entry)

        access_secret_label = canvas.create_text(420, 30, text="Access Secret:", fill="white")
        self.access_secret_entry = tk.Entry(self.master, width=30)
        canvas.create_window(500, 75, window=self.access_secret_entry)

        home_dir = os.path.expanduser("~")
        access_key_path = os.path.join(home_dir, '.secret.json')
        self.secret_file_path = access_key_path
        if os.path.exists(access_key_path):
            with open(access_key_path, 'r') as file:
                data = json.load(file)
                self.access_key = data.get("accessKey", None)
                self.access_secret = data.get("accessSecret", None)
                self.access_key_entry.insert(0, self.access_key)
                self.access_secret_entry.insert(0, self.access_secret)

        self.upload_button = tk.Button(self.master, text="请上传完税证明PDF文件", command=self.upload_file, font=("Helvetica", 14),
                            width=30, height=3)
        self.upload_button.bind("<Enter>", lambda e: self.on_enter(self.upload_button))
        self.upload_button.bind("<Leave>", lambda e: self.on_leave(self.upload_button))
        canvas.create_window(285, 250, anchor=tk.NW, window=self.upload_button)

        self.confirm_button = tk.Button(self.master, text="确认", command=self.confirm_upload, font=("Helvetica", 14),
                            width=30, height=3)
        self.confirm_button.bind("<Enter>", lambda e: self.on_enter(self.confirm_button))
        self.confirm_button.bind("<Leave>", lambda e: self.on_leave(self.confirm_button))
        canvas.create_window(285, 350, anchor=tk.NW, window=self.confirm_button)

    def create_widgets(self, frame):
        self.upload_button = tk.Button(frame, text="请上传完税证明PDF文件", command=self.upload_file,
                                       font=("Helvetica", 14),
                                       width=30, height=3)
        self.upload_button.pack(pady=20)
        self.upload_button.bind("<Enter>", lambda e: self.on_enter(self.upload_button))
        self.upload_button.bind("<Leave>", lambda e: self.on_leave(self.upload_button))

        self.confirm_button = tk.Button(frame, text="确认", command=self.confirm_upload, font=("Helvetica", 14),
                                        width=30, height=3)
        self.confirm_button.pack(pady=20)
        self.confirm_button.bind("<Enter>", lambda e: self.on_enter(self.confirm_button))
        self.confirm_button.bind("<Leave>", lambda e: self.on_leave(self.confirm_button))

    def upload_file(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if self.file_path:
            self.upload_button.config(text=self.file_path.split('/')[-1])

    def confirm_upload(self):
        if self.doing:
            messagebox.showwarning(message="有任务正在执行...")
            return

        self.doing = True
        # if self.test:
        #     self.file_path = "19年完税证明.pdf"
        #     with open("output.json", 'r') as file:
        #         invoice_dict = json.load(file)
        #         file_name = os.path.basename(self.file_path)
        #         sheet_name, _ = os.path.splitext(file_name)
        #         invoice_to_excel(invoice_dict, self.file_path, sheet_name)

        self.access_key = self.access_key_entry.get().strip()
        self.access_secret = self.access_secret_entry.get().strip()
        if self.file_path is None:
            messagebox.showwarning("没有选择文件", message="请先上传一个完税证明PDF文件")
            self.doing = False
            return
        if self.access_key is None or len(self.access_key) == 0:
            messagebox.showwarning("accessKey为空", message="请先输入完整的access key")
            self.doing = False
            return
        if self.access_secret is None or len(self.access_secret) == 0:
            messagebox.showwarning("accessSecret为空", message="请先输入完整的access secret")
            self.doing = False
            return

        self.parse_pdf()
        new_file_path = self.file_path[:-len("pdf")] + "xlsx"
        if self.failed == 0:
            messagebox.showinfo("success", message=f"PDF转Excel成功，文件路径为: {new_file_path}")
        else:
            messagebox.showwarning("failed", message=f"部分成功，有{self.failed}张图片转Excel行数据失败，图片保存在文件目录下")

        self.doing = False
        self.upload_button.config(text="请上传完税证明PDF文件")

    def parse_pdf(self):
        converted_images_folder = 'converted_images/'
        os.makedirs(converted_images_folder, exist_ok=True)

        if self.test:
            image_path = os.path.join(converted_images_folder, "invoice.jpg")
            file_name = os.path.basename(self.file_path)
            sheet_name, _ = os.path.splitext(file_name)
            ocr_manager = OCR(self.access_key, self.access_secret)
            ocr_manager.recognize(image_path, self.file_path, sheet_name)
            return

        if self.file_path.endswith('.pdf'):
            file_name = os.path.basename(self.file_path)
            sheet_name, _ = os.path.splitext(file_name)
            images = convert_from_path(self.file_path, fmt='jpeg')
            for i, image in enumerate(images):
                image_path = os.path.join(converted_images_folder, f'invoice-{i}-{sheet_name}.jpg')
                image.save(image_path, 'JPEG')

                ocr_manager = OCR(self.access_key, self.access_secret)
                if ocr_manager.recognize(image_path, self.file_path, sheet_name) < 0:
                    self.failed += 1

                os.remove(image_path)

        with open(self.secret_file_path, 'w') as file:
            data = {
                "accessKey": self.access_key,
                "accessSecret": self.access_secret
            }
            json.dump(data, file, ensure_ascii=False, indent=4)

        shutil.rmtree(converted_images_folder)

    def on_enter(self, widget):
        widget['background'] = 'grey'

    def on_leave(self, widget):
        widget['background'] = 'white'


if __name__ == '__main__':
    root = tk.Tk()
    root.title("完税证明转Excel客户端")
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    window_width = 800
    window_height = 600
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2
    root.geometry(f"{window_width}x{window_height}+{x}+{y}")
    bg_image = Image.open("images/background.jpg")  # 背景图片路径
    bg_image = bg_image.resize((window_width, window_height), Image.ANTIALIAS)
    bg_photo = ImageTk.PhotoImage(bg_image)
    bg_label = tk.Label(root, image=bg_photo)
    bg_label.place(x=0, y=0, relwidth=1, relheight=1)

    app = Application(master=root, test=False)
    app.mainloop()
